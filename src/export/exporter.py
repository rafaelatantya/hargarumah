"""Data exporter — exports property listings to JSON, CSV, and Excel formats."""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config.settings import DATA_DIR

logger = logging.getLogger(__name__)

EXPORTS_DIR = DATA_DIR / "exports"


def _ensure_dir(path: Path) -> None:
    """Ensure directory exists."""
    path.mkdir(parents=True, exist_ok=True)


def _generate_filename(prefix: str, extension: str) -> str:
    """Generate a timestamped filename."""
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"


def _build_metadata(
    listings: list[dict],
    center_lat: float,
    center_lng: float,
    radius_km: float,
) -> dict:
    """Build metadata dict for export headers."""
    sources = list({listing.get("source", "unknown") for listing in listings})
    return {
        "scraped_at": datetime.now(timezone.utc).isoformat(),
        "center_lat": center_lat,
        "center_lng": center_lng,
        "radius_km": radius_km,
        "total_listings": len(listings),
        "sources": sorted(sources),
    }


def export_json(
    listings: list[dict],
    center_lat: float,
    center_lng: float,
    radius_km: float,
    output_dir: Path | None = None,
) -> Path:
    """Export listings to JSON format.

    Args:
        listings: List of property listing dicts.
        center_lat: Search center latitude.
        center_lng: Search center longitude.
        radius_km: Search radius in km.
        output_dir: Output directory (defaults to data/exports/).

    Returns:
        Path to the created JSON file.
    """
    out_dir = output_dir or EXPORTS_DIR
    _ensure_dir(out_dir)

    filename = _generate_filename("hargarumah", "json")
    filepath = out_dir / filename

    output = {
        "metadata": _build_metadata(listings, center_lat, center_lng, radius_km),
        "listings": listings,
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    logger.info("Exported %d listings to JSON: %s", len(listings), filepath)
    return filepath


def export_csv(
    listings: list[dict],
    center_lat: float,
    center_lng: float,
    radius_km: float,
    output_dir: Path | None = None,
) -> Path:
    """Export listings to CSV format.

    Args:
        listings: List of property listing dicts.
        center_lat: Search center latitude.
        center_lng: Search center longitude.
        radius_km: Search radius in km.
        output_dir: Output directory (defaults to data/exports/).

    Returns:
        Path to the created CSV file.
    """
    out_dir = output_dir or EXPORTS_DIR
    _ensure_dir(out_dir)

    filename = _generate_filename("hargarumah", "csv")
    filepath = out_dir / filename

    if not listings:
        logger.warning("No listings to export to CSV")
        return filepath

    # Flatten images list to comma-separated string
    flat_listings = []
    for listing in listings:
        flat = dict(listing)
        if isinstance(flat.get("images"), list):
            flat["images"] = "; ".join(flat["images"])
        flat_listings.append(flat)

    fieldnames = list(flat_listings[0].keys())

    with open(filepath, "w", encoding="utf-8", newline="") as f:
        # Write metadata as comments
        meta = _build_metadata(listings, center_lat, center_lng, radius_km)
        f.write(f"# HargaRumah Export — {meta['scraped_at']}\n")
        f.write(f"# Center: ({center_lat}, {center_lng}) | Radius: {radius_km} km\n")
        f.write(f"# Total: {len(listings)} listings from {', '.join(meta['sources'])}\n")

        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(flat_listings)

    logger.info("Exported %d listings to CSV: %s", len(listings), filepath)
    return filepath


def export_xlsx(
    listings: list[dict],
    center_lat: float,
    center_lng: float,
    radius_km: float,
    output_dir: Path | None = None,
) -> Path:
    """Export listings to Excel (XLSX) format with formatting.

    Creates three sheets:
    - Listings: All data with headers and formatting
    - Summary: Metadata and statistics
    - Raw: Unformatted data for programmatic use

    Args:
        listings: List of property listing dicts.
        center_lat: Search center latitude.
        center_lng: Search center longitude.
        radius_km: Search radius in km.
        output_dir: Output directory (defaults to data/exports/).

    Returns:
        Path to the created XLSX file.
    """
    out_dir = output_dir or EXPORTS_DIR
    _ensure_dir(out_dir)

    filename = _generate_filename("hargarumah", "xlsx")
    filepath = out_dir / filename

    wb = Workbook()

    # --- Sheet 1: Listings (formatted) ---
    ws_listings = wb.active
    ws_listings.title = "Listings"  # type: ignore[union-attr]

    if listings:
        headers = list(listings[0].keys())

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws_listings.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, listing in enumerate(listings, 2):
            for col_idx, header in enumerate(headers, 1):
                value = listing.get(header, "")
                if isinstance(value, list):
                    value = "; ".join(str(v) for v in value)
                ws_listings.cell(row=row_idx, column=col_idx, value=value)  # type: ignore[union-attr]

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(len(listings) + 2, 50)):
                cell_val = ws_listings.cell(row=row_idx, column=col_idx).value  # type: ignore[union-attr]
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), 50))
            ws_listings.column_dimensions[get_column_letter(col_idx)].width = max_len + 2  # type: ignore[union-attr]

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet("Summary")
    meta = _build_metadata(listings, center_lat, center_lng, radius_km)

    summary_data = [
        ("HargaRumah Export Summary", ""),
        ("", ""),
        ("Scraped At", meta["scraped_at"]),
        ("Center Latitude", center_lat),
        ("Center Longitude", center_lng),
        ("Radius (km)", radius_km),
        ("Total Listings", len(listings)),
        ("Sources", ", ".join(meta["sources"])),
    ]

    # Price statistics
    prices = [l.get("price_idr", 0) for l in listings if l.get("price_idr")]
    if prices:
        summary_data.extend([
            ("", ""),
            ("Price Statistics", ""),
            ("Min Price (IDR)", min(prices)),
            ("Max Price (IDR)", max(prices)),
            ("Avg Price (IDR)", int(sum(prices) / len(prices))),
            ("Median Price (IDR)", sorted(prices)[len(prices) // 2]),
        ])

    title_font = Font(bold=True, size=14)
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = (
            title_font if row_idx == 1 else Font(bold=True)
        )
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40

    # --- Sheet 3: Raw (unformatted) ---
    ws_raw = wb.create_sheet("Raw")
    if listings:
        headers = list(listings[0].keys())
        for col_idx, header in enumerate(headers, 1):
            ws_raw.cell(row=1, column=col_idx, value=header)
        for row_idx, listing in enumerate(listings, 2):
            for col_idx, header in enumerate(headers, 1):
                value = listing.get(header, "")
                if isinstance(value, list):
                    value = json.dumps(value)
                ws_raw.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)
    logger.info("Exported %d listings to XLSX: %s", len(listings), filepath)
    return filepath


async def export_all(
    listings: list[dict],
    center_lat: float,
    center_lng: float,
    radius_km: float,
    output_dir: Path | None = None,
) -> dict[str, Path]:
    """Export listings to all supported formats (JSON, CSV, XLSX).

    Args:
        listings: List of property listing dicts.
        center_lat: Search center latitude.
        center_lng: Search center longitude.
        radius_km: Search radius in km.
        output_dir: Output directory (defaults to data/exports/).

    Returns:
        Dictionary mapping format name to output file path.
    """
    results = {
        "json": export_json(listings, center_lat, center_lng, radius_km, output_dir),
        "csv": export_csv(listings, center_lat, center_lng, radius_km, output_dir),
        "xlsx": export_xlsx(listings, center_lat, center_lng, radius_km, output_dir),
    }
    logger.info("Exported to all formats: %s", {k: str(v) for k, v in results.items()})
    return results
